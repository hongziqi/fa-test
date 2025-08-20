# conftest.py
import os
import time
import pandas as pd
import pytest
from openpyxl import load_workbook

# from test_fa_fwd_npu_ac import test_results, RESULT_DIR  # 导入你的全局变量

from test_fa_fwd_npu_prof_tune import test_results, RESULT_DIR  # 导入你的全局变量


def pytest_addoption(parser):
    """添加命令行选项以支持分批测试"""
    parser.addoption("--start-index", action="store", default=0, type=int, 
                    help="测试案例的起始索引")
    parser.addoption("--batch-size", action="store", default=5, type=int,
                    help="每批运行的测试案例数量")


def pytest_sessionfinish(session, exitstatus):
    """
    测试会话结束时，保存测试结果到文件
    """
    if not test_results:
        print(">> No test results to save. `test_results` is empty.")
        return
    # try:
    #     timestamp = time.strftime("%Y%m%d_%H%M%S")
    #     # timestamp = ""
    #     result_file = os.path.join(RESULT_DIR, f"test_results_step64_01_{timestamp}.xlsx")
    #     # result_file = os.path.join(RESULT_DIR, f"test_results_prof_01_{timestamp}.xlsx")
    #     df = pd.DataFrame(test_results).fillna("")
    #     df.to_excel(result_file, index=False)
    #     print(f"\n>> 测试完成，结果已保存至 {result_file}")
    #     print(f"总计 {len(df)} 个测试用例")
    #     print(f"通过: {len(df[df['result'] == 'Success'])}")
    #     print(f"异常: {len(df[df['result'] == 'Error'])}")
    # except Exception as e:
    #     print(f"保存测试结果时发生错误: {e}")
    #     pytest.fail(f"测试结果保存失败: {e}")
    try:
        result_file = os.path.join(RESULT_DIR, "test_results_npu_tune_batch.xlsx")
        df = pd.DataFrame(test_results).fillna("")

        if not os.path.exists(result_file):
            # 文件不存在 → 直接写新文件
            df.to_excel(result_file, index=False)
        else:
            # 文件已存在 → 追加写入
            book = load_workbook(result_file)
            with pd.ExcelWriter(result_file, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                startrow = writer.sheets["Sheet1"].max_row
                df.to_excel(writer, index=False, header=False, startrow=startrow)

        print(f"\n>> 当前批次结果已追加至 {result_file}")
    except Exception as e:
        print(f"保存测试结果时发生错误: {e}")
        pytest.fail(f"测试结果保存失败: {e}")
